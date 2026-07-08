from datetime import datetime
import os
import sys
from pydoc import text

from openpyxl.formatting import rule
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D

class ExcelHandler:
    """Generator for ZPL (Test Report) Excel workbooks.

    Manages workbook creation, sheet formatting, conditional coloring,
    and data population following the standard ZPL report layout.
    Uses openpyxl under the hood.
    """

    ##TODO sposta tutto in un file di configurazione e importalo qui
    THIN_STYLE = Side(style="thin", color="FF0000")
    BORDER_STYLE = Border(left=THIN_STYLE, right=THIN_STYLE, top=THIN_STYLE, bottom=THIN_STYLE)
    BIGGER_BOLD_FONT = Font(size=14, bold=True)
    BIG_BOLD_FONT = Font(size=13, bold=True)
    BIGGER_FONT = Font(size=14)
    NORMAL_FONT = Font(size=12)
    ITALIC_FONT = Font(size=12, italic=True)
    ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ZPL_QUERY_FILL = ""
    DATE_QUERY_FILL = ""

    #Stili grafici per le celle come colori sfondo, scrite e bordi
    #colore di sfondo
    PASS_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    BLOCKED_FILL = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
    
    FULL_RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    FULL_GREEN_FILL = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
    #colore della scritta
    PASS_FONT = Font(color="FF009700")
    FAIL_FONT = Font(color="FF9C0006")
    BLOCKED_FONT = Font(color="FF9C6500")

    #bordi delle celle
    LINE_SIDE = Side(style="thin", color="FF000000")
    DOTTED_SIDE = Side(style="dotted", color="FF000000")

    def __init__(self):
        """Initialize a new workbook with a single active sheet."""
        self.workbook = Workbook()
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        self.sheet : Worksheet | None = None


    def fill_cell(self, cell_number:str, value: str, sheet: Worksheet | None = None,
                  font:Font = Font(), alignment:Alignment = Alignment(),
                  fill:PatternFill = PatternFill()):
        """Write a value into a single cell and apply optional styling.

        Args:
            cell_number: Cell reference (e.g. "A1", "C12").
            value: Content to write into the cell.
            sheet: Target worksheet (defaults to the active sheet).
            font: Font style to apply.
            alignment: Alignment to apply.
            fill: Background fill to apply.

        Raises:
            ValueError: If cell_number is empty.
        """
        if sheet is None:
            if self.sheet is None:
                raise ValueError("No active sheet available.")
            sheet = self.sheet
        
        if not cell_number:
            raise ValueError("Cell reference cannot be empty.")
        cell = sheet[cell_number]
        cell.value = value
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill


    def merge_and_fill(self, cell_range: str, value: str, sheet: Worksheet | None = None,
                       font:Font = Font(), alignment:Alignment = Alignment(),
                       fill:PatternFill = PatternFill()):
        """Merge a range of cells and write a value into the top-left cell.

        Args:
            cell_range: Range to merge (e.g. "B2:D2").
            value: Content to write into the merged cell.
            sheet: Target worksheet (defaults to the active sheet).
            font: Font style to apply.
            alignment: Alignment to apply.
            fill: Background fill to apply.
        """
        if sheet is None:
            if self.sheet is None:
                raise ValueError("No active sheet available.")
            sheet = self.sheet

        sheet.merge_cells(cell_range)
        cell_coordinate = cell_range.split(':')[0]
        
        self.fill_cell(cell_coordinate, value, sheet, font, alignment, fill)


    def set_common_header(self, sheet: Worksheet | None = None):
        """Populate the standard ZPL report header (rows 1-10).

        Includes the company logo, report title, SAP OLE link fields,
        test performer info, product details, and conformance status.

        Args:
            sheet: Target worksheet (defaults to the active sheet).
        """
        if sheet is None:
            sheet = self.sheet
        
        # Titolo principale
        self.merge_and_fill('C1:E1', "RAPPORTO DI PROVA\nTest Report", sheet,
                            self.BIGGER_FONT, self.ALIGNMENT_CENTER)

        # Data e pagina
        
        self.fill_cell('F1', "Data/Date", sheet, alignment=self.ALIGNMENT_CENTER)
        self.fill_cell('G1', "=SAP.OLELinkServer.ItemObject.1|SAPOLELinkServerDMS!'!Char_Value.DRAW.017.PLCC008.1'", sheet, alignment=self.ALIGNMENT_CENTER)

        self.fill_cell('H1', "Pag/Page", sheet, alignment=self.ALIGNMENT_CENTER)
        self.fill_cell('I1', "5/6", sheet, alignment=self.ALIGNMENT_CENTER)
        
        self.sheet.row_dimensions[1].height = 36

        # Inserire immagine
        # ? create a function for putting image on sheet
        base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        img = Image(os.path.join(base, "export", "src", "logo.jpg") if hasattr(sys, '_MEIPASS') else os.path.join(os.path.dirname(__file__), "src", "logo.jpg"))
        
        marker = AnchorMarker(
            col=1,  # colonna B (indice base 0)
            row=0,  # riga 1 (indice base 0)
            colOff=pixels_to_EMU(5),   # padding sinistro 5 px
            rowOff=pixels_to_EMU(5)    # padding superiore 5 px
        )
        img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height)))
        sheet.add_image(img)

        # Laboratorio
        self.fill_cell('B2', "Laboratorio Prove", sheet, font=self.BIGGER_FONT)
        self.fill_cell('B3', "Test Laboratory", sheet, font=self.ITALIC_FONT)
        # Numero ZPL
        self.merge_and_fill('C2:D3', "Numero/Number:", sheet, alignment=self.ALIGNMENT_LEFT)
        self.merge_and_fill('E2:I3', "=SAP.OLELinkServer.ItemObject.1|SAPOLELinkServerDMS!'!Document.DRAW.DOKNR'", sheet, font=Font(bold=True), alignment=self.ALIGNMENT_LEFT)

        # Info test
        self.merge_and_fill('B5:C5', "Esecutore del test / Test performer", sheet)
        self.merge_and_fill('D5:I5', "Inserire nome", sheet)

        self.merge_and_fill('B6:C6', "Oggetto di prova/Test Object", sheet)
        self.merge_and_fill('D6:I6', "inserire Gwc0de", sheet)

        self.merge_and_fill('B7:C7', "Nome prodotto", sheet)
        self.merge_and_fill('D7:I7', "=SAP.OLELinkServer.ItemObject.1|SAPOLELinkServerDMS!'!Char_Value.DRAW.017.PLCC016.1'", sheet)

        self.merge_and_fill('B8:C8', "Firmware Version", sheet)
        self.merge_and_fill('D8:I8', "inserire FW", sheet)

        # Conforme
        self.merge_and_fill('B9:C9', "Conforme/Conformance", sheet)
        self.merge_and_fill('D9:I9', "=SAP.OLELinkServer.ItemObject.1|SAPOLELinkServerDMS!'!Char_Value.DRAW.017.PLCC011.1'", sheet)
        self.color_conform_cell('D9', sheet)

        # Nota
        self.merge_and_fill('D10:I10', "Inserire i dispositivi e i rispettivi fw dei prodotti utilizzati", sheet)

        # System info
        self.merge_and_fill('B10:C10', "System info & Configuration", sheet)

        self.set_borders("B1", "I3", sheet=sheet, border_style_in=self.LINE_SIDE, border_style_out=self.LINE_SIDE)
        self.set_borders("B5", "I10", sheet=sheet, border_style_in=self.LINE_SIDE, border_style_out=self.LINE_SIDE)


    def set_datasheet(self, sheet: Worksheet | None = None, start_row: int = 12, section: str = "", table_titles: list[str] =  [], data: list[tuple] = ()):
        """Write a data table with section title, column headers, and rows.

        Automatically handles column merging when the number of titles is less
        than the available columns (B through I), and applies border formatting.

        Args:
            sheet: Target worksheet (defaults to the active sheet).
            start_row: Row number where the section begins.
            section: Section title displayed above the table.
            table_titles: List of column header labels.
            data: List of tuples, each representing a row of data values.
        """
        if sheet is None:
            sheet = self.sheet

        # Implement datasheet drawing and filling logic here
        # set section title
        if section:
            #self.sheet[f"B{start_row}"] = section
            self.merge_and_fill(f"B{start_row}:I{start_row}", section, sheet, font=self.BIG_BOLD_FONT)
        
        # set titles
        if len(table_titles) < 8:
                start_merge_col = chr(66 + len(table_titles) - 1)
                self.merge_and_fill(f"{start_merge_col}{start_row + 1}:I{start_row + 1}","",sheet=sheet)
                
        for offset, title in enumerate(table_titles):  
            self.sheet[f"{chr(66 + offset)}{start_row + 1}"] = title
 

        # set data
        for offset, row_data in enumerate(data):
            for col_offset, value in enumerate(row_data):
                if len(row_data) > len(table_titles):
                    raise ValueError(f"Data row has more columns than titles. Row data: {row_data}, Titles: {table_titles}")
                self.sheet[f"{chr(66 + col_offset)}{start_row + 2 + offset}"] = value
        #unire le ultime celle per rendere la tabella una B:I
        if len(table_titles) < 8: ##TODO cambiare 8 e I in magic number
            start_merge_col = chr(66 + len(table_titles) - 1)
            for row_idx in range(start_row + 2, start_row + 2 + len(data)):
                self.merge_and_fill(f"{start_merge_col}{row_idx}:I{row_idx}","",sheet=sheet)       

            
        
        self.set_borders(f'B{start_row + 1}', f'I{len(data) + 1 + start_row}', sheet=sheet, border_style_in=self.DOTTED_SIDE, border_style_out=self.LINE_SIDE)
        

    def make_new_sheet(self, sheet_name: str):
        """Create a new worksheet and set it as the active target for subsequent operations.

        Args:
            sheet_name: Name for the new sheet tab.
        """
        self.sheet = self.workbook.create_sheet(title=sheet_name)
        

    def save(self, filename: str = "", parent_path: str = ""):
        """Save the workbook to disk with a timestamped filename.

        The final filename has the format: <filename>_YYYYMMDD_HHMMSS.xlsx

        Args:
            filename: Base filename (without timestamp). Defaults to "report_zpl".
            parent_path: Directory where the file will be saved. Defaults to cwd.

        Returns:
            The absolute path of the saved .xlsx file.
        """
        if not parent_path:
            parent_path = os.getcwd()
        if not filename:
            filename = f"report_zpl"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        
        filename = filename.replace(".xlsx", f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        full_path = os.path.join(parent_path, filename)
        if not os.path.exists(parent_path):
            os.makedirs(parent_path)
        print(f"Saving Excel file to: {full_path}")
        self.workbook.save(full_path)
        return full_path

    def set_borders (self, cell_start: str, cell_end: str, sheet: Worksheet | None = None, border_style_in: Side = LINE_SIDE, border_style_out: Side = LINE_SIDE):
        """Apply border formatting to a rectangular cell range.

        Differentiates between inner borders and outer (perimeter) borders.

        Args:
            cell_start: Top-left cell reference (e.g. "B1").
            cell_end: Bottom-right cell reference (e.g. "I10").
            sheet: Target worksheet (defaults to the active sheet).
            border_style_in: Side style for internal cell borders.
            border_style_out: Side style for the outer perimeter.
        """
        if sheet is None:
            sheet = self.sheet

        cell_range = f"{cell_start}:{cell_end}"
        
        min_col= ord(cell_start[0]) - ord("A") + 1
        min_row= int(cell_start[1:])
        max_col = ord(cell_end[0]) - ord("A") + 1
        max_row = int(cell_end[1:])

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell=sheet.cell(row=row, column=col)

                #bordi interni
                left = border_style_in
                right = border_style_in
                top = border_style_in
                bottom = border_style_in

                # Bordi esterni
                if row == min_row:
                    top = border_style_out
                if row == max_row:
                    bottom = border_style_out
                if col == min_col:
                    left = border_style_out
                if col == max_col :
                    right = border_style_out

                cell.border = Border(
                    left=left,
                    right=right,
                    top=top,
                    bottom=bottom
                )

    def color_cells(self, cell_start:str, cell_end:str, value:str, fill:PatternFill, font:Font, sheet: Worksheet | None = None):
        """Add a conditional formatting rule that highlights cells containing a value.

        Uses ISNUMBER(SEARCH(...)) so the match is case-insensitive and partial.

        Args:
            cell_start: First cell of the range.
            cell_end: Last cell of the range.
            value: Text to search for inside cells.
            fill: Background fill to apply when the condition is met.
            font: Font style to apply when the condition is met.
            sheet: Target worksheet (defaults to the active sheet).
        """
        if sheet is None:
            sheet = self.sheet

        cell_range = f"{cell_start}:{cell_end}"
        sheet.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{value}",{cell_start}))'],
                fill=fill,
                font=font
            )
        )

    def color_state(self, column:str, sheet: Worksheet | None = None):
        """Apply pass/fail/blocked conditional coloring to an entire column.

        Adds three rules (pass=green, fail=red, blocked=yellow) from row 13
        to the last used row.

        Args:
            column: Column letter (e.g. "E", "F").
            sheet: Target worksheet (defaults to the active sheet).
        """
        if sheet is None:
            sheet = self.sheet

        cell_range = f"{column}13:{column}{sheet.max_row}"

        # PASS
        self.color_cells(cell_start=f"{column}13", cell_end=f"{column}{sheet.max_row}", value="pass", fill=self.PASS_FILL, font=self.PASS_FONT, sheet=sheet)

        # FAIL
        self.color_cells(cell_start=f"{column}13", cell_end=f"{column}{sheet.max_row}", value="fail", fill=self.FAIL_FILL, font=self.FAIL_FONT, sheet=sheet)

        # BLOCKED
        self.color_cells(cell_start=f"{column}13", cell_end=f"{column}{sheet.max_row}", value="BLOCKED", fill=self.BLOCKED_FILL, font=self.BLOCKED_FONT, sheet=sheet)

    def color_conform_cell(self, cell:str, sheet: Worksheet | None = None):
        """Apply conformance coloring to a single cell (green=SI, red=NO).

        Args:
            cell: Cell reference (e.g. "D9").
            sheet: Target worksheet (defaults to the active sheet).
        """
        if sheet is None:
            sheet = self.sheet

        # CONFORME
        self.color_cells(cell_start=cell, cell_end=cell, value="SI", fill=self.FULL_GREEN_FILL,font=self.NORMAL_FONT, sheet=sheet)

        # NON CONFORME
        self.color_cells(cell_start=cell, cell_end=cell, value="NO", fill=self.FULL_RED_FILL,font=self.NORMAL_FONT, sheet=sheet)

#
    #def color_state (self, column:str, sheet: Worksheet | None = None):
    #    if sheet is None:
    #         sheet = self.sheet
#
    #    #col= ord(column)-64
    #    cell_range = f"{column}13:{column}{sheet.max_row}"
#
    #       # PASS
    #    sheet.conditional_formatting.add(
    #        cell_range,
    #        FormulaRule(
    #            formula=[f'ISNUMBER(SEARCH("pass",${column}13))'],
    #            fill=self.PASS_FILL,
    #            font=self.PASS_FONT
    #        )
    #    )
    #
    #    # FAIL
    #    sheet.conditional_formatting.add(
    #        cell_range,
    #        FormulaRule(
    #            formula=[f'ISNUMBER(SEARCH("fail",${column}13))'],
    #            fill=self.FAIL_FILL,
    #            font=self.FAIL_FONT
    #        )
    #    )
    #
    #    # BLOCKED
    #    sheet.conditional_formatting.add(
    #        cell_range,
    #        FormulaRule(
    #            formula=[f'ISNUMBER(SEARCH("BLOCKED",${column}13))'],
    #            fill=self.BLOCKED_FILL,
    #            font=self.BLOCKED_FONT
    #        )
    #    )
#

if __name__ == "__main__":
    excel_handler = ExcelHandler()
    excel_handler.set_common_header()
    excel_handler.set_datasheet(section = "ELENCO ISSUE E BUG / ISSUE AND BUG LIST",
        table_titles=[        
        "TYPE",
        "ID SEGNALAZIONE",
        "TITOLO",
        "ID TEST FAIL",
        "NOTE"],
        data=[("Issue", "456789", "Titolo 1", "ID Test 1"),
              ("Bug", "123456", "Titolo 2")])
    excel_handler.save("test_report.xlsx")
